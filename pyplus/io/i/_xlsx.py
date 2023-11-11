import typing
import datetime
import itertools
import contextlib

import openpyxl

M = typing.TypeVar("M")

_Book: typing.TypeAlias = openpyxl.workbook.workbook.Workbook
_Sheet: typing.TypeAlias = openpyxl.worksheet.worksheet.Worksheet
_Value: typing.TypeAlias = str | float | datetime.datetime | None


def _get_sheet(wb: _Book, index: int | None) -> _Sheet | None:
    if index is None:
        return typing.cast(_Sheet, wb.active)

    return wb.worksheets[index]


def _read_lines(
    fpath: str,
    *,
    skiprows: int = 0,
    index: int | None = None,
    read_only: bool = True,
) -> typing.Iterator[list[_Value]]:
    with contextlib.closing(
        openpyxl.load_workbook(filename=fpath, read_only=read_only)
    ) as wb:
        ws = _get_sheet(wb, index)
        if ws is None:
            return

        reader: typing.Iterator[typing.Iterable[_Value]]
        reader = ws.iter_rows(values_only=True)
        if skiprows:
            reader = itertools.islice(reader, skiprows, None)

        yield from map(list, reader)


def _read_models(
    fpath: str,
    *,
    header: int = 0,
    index: int | None = None,
    read_only: bool = True,
    converter: typing.Callable[[dict[str, _Value]], M],
) -> typing.Iterator[M]:
    with contextlib.closing(
        openpyxl.load_workbook(filename=fpath, read_only=read_only)
    ) as wb:
        ws = _get_sheet(wb, index)
        if ws is None:
            return

        reader: typing.Iterator[typing.Iterable[_Value]]
        reader = ws.iter_rows(values_only=True)
        if header:
            reader = itertools.islice(reader, header, None)

        keys: list[str]
        for i, row in enumerate(reader):
            values: list[_Value] = list(row)
            if i == 0:
                keys = typing.cast(list[str], values)
                continue

            data: dict[str, _Value]
            data = {
                key: values[i] if i < len(values) else None
                for i, key in enumerate(keys)
            }
            yield converter(data)
